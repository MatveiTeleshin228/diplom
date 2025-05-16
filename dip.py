import sys
import re
import logging
import time
import openpyxl
import psycopg2
from psycopg2 import sql
from PySide6.QtCore import (
    Qt,
    QAbstractTableModel,
    QModelIndex,
    QThreadPool,
    QRunnable,
    QObject,
    Signal,
    QPropertyAnimation,
    QEasingCurve,
    QRegularExpression,
    QSortFilterProxyModel,
)
from PySide6.QtGui import QFont, QAction, QRegularExpressionValidator
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QTableView,
    QPushButton,
    QLineEdit,
    QLabel,
    QMessageBox,
    QDialog,
    QFormLayout,
    QComboBox,
    QTabWidget,
    QProgressDialog,
    QHeaderView,
    QGraphicsOpacityEffect,
    QSpinBox,
    QSizePolicy,
    QPlainTextEdit,
    QFrame,
    QCheckBox,
)
from PySide6.QtWidgets import QDialogButtonBox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Настройка логирования
logging.basicConfig(
    level=logging.DEBUG,
    filename="app.log",
    filemode="a",
    format="%(asctime)s %(levelname)s: %(message)s",
)


def exception_hook(exc_type, exc_value, exc_traceback):
    logging.error(
        "Неперехваченное исключение", exc_info=(exc_type, exc_value, exc_traceback)
    )
    sys.exit(1)


sys.excepthook = exception_hook


class Database:
    def __init__(self):
        self.connection = None
        self.connect()

    def connect(self):
        try:
            self.connection = psycopg2.connect(
                dbname="postgres",
                user="postgres.okwxukyvivltgexgjlaf",
                password="pqiBU3JavEhOxm6E",
                host="aws-0-eu-north-1.pooler.supabase.com",
                port="6543",
                connect_timeout=5,
            )
            self.connection.autocommit = False
            logging.info("Успешное подключение к базе данных")
        except Exception as e:
            logging.error(f"Ошибка подключения к базе данных: {e}")
            raise

    def execute_query(self, query, params=None, fetch=False):
        try:
            if not self.connection or self.connection.closed:
                self.connect()

            with self.connection.cursor() as cursor:
                cursor.execute(query, params)

                if fetch:
                    result = cursor.fetchall()
                    columns = (
                        [desc[0] for desc in cursor.description]
                        if cursor.description
                        else []
                    )
                    self.connection.commit()
                    return result, columns
                else:
                    self.connection.commit()
                    return True

        except Exception as e:
            if self.connection:
                self.connection.rollback()
            logging.error(f"Ошибка выполнения запроса: {e}")
            return False

    def close(self):
        if self.connection:
            self.connection.close()
            logging.info("Соединение с базой данных закрыто")


db = Database()

# QSS-стилизация приложения
APP_QSS = """
QMainWindow {
    background-color: #f0f0f0;
}

QTabWidget::pane {
    border: 1px solid #C2C7CB;
    background: #ffffff;
}

QTabBar::tab {
    background: #e0e0e0;
    padding: 10px;
    margin: 2px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
}
QTabBar::tab:selected {
    background: #ffffff;
    border-bottom: 2px solid #3498db;
}

QPushButton {
    background-color: #3498db;
    color: white;
    border: none;
    padding: 8px 16px;
    border-radius: 4px;
}
QPushButton:hover {
    background-color: #2980b9;
}
QPushButton:pressed {
    background-color: #1c5980;
}

QLineEdit, QComboBox, QSpinBox {
    border: 1px solid #bdc3c7;
    border-radius: 4px;
    padding: 4px;
    background-color: #ffffff;
}
QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
    border: 1px solid #3498db;
}

QTableView::item:selected {
    background: #3498db;  /* Ярко-синий */
    color: white;         /* Белый текст */
    border: none;
}
QHeaderView::section {
    background-color: #ecf0f1;
    padding: 4px;
    border: 1px solid #bdc3c7;
}
QPushButton[text="Одобрить заявку"] {
    background-color: #27ae60;
}
QPushButton[text="Одобрить заявку"]:hover {
    background-color: #219955;
}
QPushButton[text="Отклонить заявку"] {
    background-color: #e74c3c;
}
QPushButton[text="Отклонить заявку"]:hover {
    background-color: #c0392b;
}
QPushButton[text="Просмотр деталей"] {
    background-color: #3498db;
}
RoomWidget {
    background-color: #f5f7fa;
}

RoomWidget QLabel#title {
    font-size: 18px;
    font-weight: bold;
    color: #2c3e50;
    padding: 10px;
}

RoomWidget QTableView {
    border: 1px solid #bdc3c7;
    border-radius: 5px;
    background-color: white;
}

RoomWidget QHeaderView::section {
    background-color: #3498db;
    color: white;
    padding: 5px;
    border: none;
}

RoomWidget QFrame#infoFrame {
    background-color: #f8f9fa;
    border: 1px solid #dee2e6;
    border-radius: 5px;
    padding: 15px;
}

RoomWidget QLabel#studentsTitle {
    font-size: 14px;
    font-weight: bold;
    color: #2c3e50;
    padding-bottom: 10px;
}

RoomWidget QLabel#studentsList {
    font-size: 13px;
    color: #495057;
    line-height: 1.4;
}
"""


# Окно просмотра логов
class LogViewerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Просмотр логов")
        self.resize(800, 600)  # Увеличим размер окна
        layout = QVBoxLayout(self)
        
        # Добавляем поясняющий текст
        info_label = QLabel("Содержимое лог-файла app.log:")
        layout.addWidget(info_label)
        
        self.text_edit = QPlainTextEdit(self)
        self.text_edit.setReadOnly(True)
        self.text_edit.setFont(QFont("Courier New", 10))  # Моноширинный шрифт для логов
        layout.addWidget(self.text_edit)
        
        button_layout = QHBoxLayout()
        refresh_button = AnimatedButton("Обновить")
        refresh_button.clicked.connect(self.load_logs)
        
        # Кнопка очистки логов
        clear_button = AnimatedButton("Очистить логи")
        clear_button.clicked.connect(self.clear_logs)
        clear_button.setStyleSheet("background-color: #e74c3c;")
        
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(clear_button)
        layout.addLayout(button_layout)
        
        self.load_logs()

    def load_logs(self):
        try:
            # Пробуем разные кодировки по очереди
            encodings = ['utf-8', 'cp1251', 'iso-8859-1', 'utf-16']
            
            for encoding in encodings:
                try:
                    with open("app.log", "r", encoding=encoding) as f:
                        content = f.read()
                        self.text_edit.setPlainText(content)
                        return
                except UnicodeDecodeError:
                    continue
                    
            # Если ни одна кодировка не подошла, читаем как бинарный файл
            with open("app.log", "rb") as f:
                content = f.read().decode('utf-8', errors='replace')
                self.text_edit.setPlainText(content)
                
        except FileNotFoundError:
            self.text_edit.setPlainText("Лог-файл не найден. Он будет создан автоматически при следующем событии логирования.")
        except Exception as e:
            self.text_edit.setPlainText(f"Ошибка чтения лог-файла: {str(e)}\n\nПопробуйте очистить логи или проверьте права доступа к файлу.")

    def clear_logs(self):
        reply = QMessageBox.question(
            self,
            "Очистка логов",
            "Вы уверены, что хотите очистить файл логов?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                with open("app.log", "w", encoding='utf-8') as f:
                    f.write("")
                self.load_logs()
                QMessageBox.information(self, "Успех", "Файл логов успешно очищен.")
            except Exception as e:
                QMessageBox.warning(self, "Ошибка", f"Не удалось очистить логи: {str(e)}")

# Модель для списка студентов
class StudentsTableModel(QAbstractTableModel):
    HEADERS = ["ID", "ФИО", "Пол", "Возраст", "Курс", "Факультет", "Телефон", "Комната"]  # Добавлен столбец "Телефон"

    def __init__(self):
        super().__init__()
        self._students = []
        self.load_data()

    def load_data(self):
        query = """
        SELECT s.id, s.fio, s.pol, s.vozrast, s.kurs, s.fakultet, s.number_phone, r.id as room_id 
        FROM students s
        LEFT JOIN rooms r ON s.room_id = r.id
        ORDER BY s.id
        """
        result, _ = db.execute_query(query, fetch=True)
        if result:
            self.beginResetModel()
            self._students = [
                {
                    "id": str(row[0]),
                    "fio": row[1],
                    "pol": row[2],
                    "vozrast": row[3],
                    "kurs": row[4],
                    "fakultet": row[5],
                    "number_phone": row[6] if row[6] else "не указан",  # Добавлено поле телефона
                    "room_id": str(row[7]) if row[7] else "еще не заселен",
                }
                for row in result
            ]
            self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self._students)

    def columnCount(self, parent=QModelIndex()):
        return len(self.HEADERS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        student = self._students[index.row()]
        if role == Qt.DisplayRole:
            return [
                student["id"],
                student["fio"],
                student["pol"],
                student["vozrast"],
                student["kurs"],
                student["fakultet"],
                student["number_phone"],  # Добавлено отображение телефона
                student["room_id"],
            ][index.column()]
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.HEADERS[section]
        return None

    def add_student(self, student):
        # Проверяем подключение
        if not db.connection or db.connection.closed:
            db.connect()

        # Обрабатываем пустой номер телефона
        number_phone = student.get("number_phone", "").strip()
        if not number_phone:
            number_phone = None  # Устанавливаем NULL вместо пустой строки

        query = """
        INSERT INTO students (fio, pol, vozrast, kurs, fakultet, number_phone) 
        VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """
        params = (
            student["fio"],
            student["pol"],
            student["vozrast"],
            student["kurs"],
            student["fakultet"],
            number_phone,  # Используем обработанное значение
        )

        try:
            # Выполняем запрос
            result = db.execute_query(query, params, fetch=True)
            
            # Проверяем результат
            if not result or not result[0]:  # Если нет результата
                return False

            new_id = result[0][0][0]  # Получаем ID
            student["id"] = str(new_id)
            student["room_id"] = "еще не заселен"  # Новый студент не заселен

            # Добавляем в модель
            row_position = len(self._students)
            self.beginInsertRows(QModelIndex(), row_position, row_position)
            self._students.append(student)
            self.endInsertRows()

            logging.info(f"Добавлен студент ID {new_id}: {student}")
            return True

        except Exception as e:
            logging.error(f"Ошибка при добавлении студента: {e}")
            if self.connection:
                self.connection.rollback()
            return False

    def update_student(self, row, student):
        if 0 <= row < len(self._students):
            query = """
            UPDATE students 
            SET fio = %s, pol = %s, vozrast = %s, kurs = %s, fakultet = %s, number_phone = %s 
            WHERE id = %s
            """
            params = (
                student["fio"],
                student["pol"],
                student["vozrast"],
                student["kurs"],
                student["fakultet"],
                student["number_phone"],  # Добавлен телефон
                int(student["id"]),
            )
            if db.execute_query(query, params):
                # Сохраняем room_id из старой записи
                student["room_id"] = self._students[row]["room_id"]
                self._students[row] = student
                self.dataChanged.emit(
                    self.index(row, 0), self.index(row, self.columnCount() - 1)
                )
                logging.info("Обновлен студент в строке %d: %s", row, student)
                return True
        return False

    def remove_student(self, row):
        if 0 <= row < len(self._students):
            student_id = int(self._students[row]["id"])
            query = "DELETE FROM students WHERE id = %s"
            if db.execute_query(query, (student_id,)):
                self.beginRemoveRows(QModelIndex(), row, row)
                student = self._students.pop(row)
                self.endRemoveRows()
                logging.info("Удалён студент: %s", student)
                return True
        return False

    def get_student(self, row):
        if 0 <= row < len(self._students):
            return self._students[row]
        return None

    def update_multiple_students(self, rows, student_data):
        """Обновляет несколько студентов"""
        if not rows:
            return False
        
        updated_ids = []
        for row in sorted(rows, reverse=True):
            if 0 <= row < len(self._students):
                student = self._students[row]
                updated_ids.append(student["id"])
                
                # Обновляем только те поля, которые были изменены
                for key in student_data:
                    if key in student:
                        student[key] = student_data[key]
        
        if not updated_ids:
            return False
            
        # Формируем SQL запрос для обновления
        set_parts = []
        params = []
        
        if "fio" in student_data:
            set_parts.append("fio = %s")
            params.append(student_data["fio"])
        if "pol" in student_data:
            set_parts.append("pol = %s")
            params.append(student_data["pol"])
        if "vozrast" in student_data:
            set_parts.append("vozrast = %s")
            params.append(student_data["vozrast"])
        if "kurs" in student_data:
            set_parts.append("kurs = %s")
            params.append(student_data["kurs"])
        if "fakultet" in student_data:
            set_parts.append("fakultet = %s")
            params.append(student_data["fakultet"])
        if "number_phone" in student_data:
            set_parts.append("number_phone = %s")
            params.append(student_data["number_phone"])
        
        if not set_parts:
            return False
            
        # Преобразуем ID студентов в целые числа
        student_ids = [int(id) for id in updated_ids]
        
        query = f"UPDATE students SET {', '.join(set_parts)} WHERE id = ANY(%s::int[])"
        params.append(student_ids)
        
        if db.execute_query(query, params):
            self.dataChanged.emit(
                self.index(min(rows), 0),
                self.index(max(rows), self.columnCount() - 1)
            )
            return True
        return False

    def remove_multiple_students(self, rows):
        """Удаляет несколько студентов"""
        if not rows:
            return False
        
        student_ids = []
        for row in sorted(rows, reverse=True):
            if 0 <= row < len(self._students):
                student_ids.append(int(self._students[row]["id"]))
        
        if not student_ids:
            return False
            
        query = "DELETE FROM students WHERE id = ANY(%s::int[])"
        if db.execute_query(query, (student_ids,)):
            # Удаляем из модели
            for row in sorted(rows, reverse=True):
                if 0 <= row < len(self._students):
                    self.beginRemoveRows(QModelIndex(), row, row)
                    self._students.pop(row)
                    self.endRemoveRows()
            return True
        return False


# Модель для списка комнат
class RoomsTableModel(QAbstractTableModel):
    HEADERS = ["ID", "Этаж", "Кол-во мест", "Свободно"]

    def __init__(self):
        super().__init__()
        self._rooms = []
        self.load_data()

    def load_data(self):
        query = "SELECT id, etazh, kol_mest, svobodno FROM rooms ORDER BY id"
        result, _ = db.execute_query(query, fetch=True)
        if result:
            self.beginResetModel()
            self._rooms = [
                {
                    "id": str(row[0]),
                    "etazh": str(row[1]),
                    "kol_mest": str(row[2]),
                    "svobodno": str(row[3]),
                }
                for row in result
            ]
            self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self._rooms)

    def columnCount(self, parent=QModelIndex()):
        return len(self.HEADERS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        room = self._rooms[index.row()]
        if role == Qt.DisplayRole:
            return [room["id"], room["etazh"], room["kol_mest"], room["svobodno"]][
                index.column()
            ]
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.HEADERS[section]
        return None

    def update_room_availability(self, room_id, change):
        """Обновляет количество свободных мест в комнате"""
        query = "UPDATE rooms SET svobodno = svobodno + %s WHERE id = %s"
        if db.execute_query(query, (change, room_id)):
            self.load_data()  # Полностью перезагружаем данные
            return True
        return False


# Модель для заявок
class RequestsTableModel(QAbstractTableModel):
    HEADERS = ["ID", "Тип заявки", "Статус", "Студент", "Комната", "Дата создания"]
    STATUS_VALUES = ["Создана", "В обработке", "Выполнена", "Отклонена"]

    def __init__(self):
        super().__init__()
        self._requests = []
        self.load_data()

    def load_data(self):
        query = """
        SELECT 
            r.id, 
            r.type, 
            r.status, 
            s.fio, 
            r.room_id, 
            (r.created_at AT TIME ZONE 'UTC' + INTERVAL '3 hours')::timestamp AS created_at
        FROM requests r
        JOIN students s ON r.student_id = s.id
        ORDER BY r.id DESC
        """
        result = db.execute_query(query, fetch=True)
        if result:
            result, _ = result
            self.beginResetModel()
            self._requests = [
                {
                    "id": str(row[0]),
                    "type": row[1],
                    "status": row[2],
                    "student_fio": row[3],
                    "room_id": str(row[4]) if row[4] else "не указана",
                    "created_at": row[5].strftime("%d.%m.%Y %H:%M") if row[5] else "не указана"
                }
                for row in result
            ]
            self.endResetModel()

    def add_request(self, request_data):
        query = """
        INSERT INTO requests (type, status, student_id, room_id) 
        VALUES (%s, %s, %s, %s) 
        RETURNING 
            id, 
            (created_at AT TIME ZONE 'UTC' + INTERVAL '3 hours')::timestamp AS created_at
        """
        params = (
            request_data["type"],
            request_data["status"],
            int(request_data["student_id"]),
            (
                int(request_data["room_id"])
                if request_data["room_id"] != "не указана"
                else None
            ),
        )
        result = db.execute_query(query, params, fetch=True)
        if result:
            request_data["id"] = str(result[0][0][0])
            request_data["created_at"] = result[0][0][1].strftime("%d.%m.%Y %H:%M")

            self._requests.insert(0, {
                "id": request_data["id"],
                "type": request_data["type"],
                "status": request_data["status"],
                "student_fio": request_data["student_fio"],
                "room_id": request_data["room_id"],
                "created_at": request_data["created_at"]
            })
            
            # Используем beginResetModel/endResetModel для обновления всей таблицы
            self.beginResetModel()
            self.endResetModel()
            
            logging.info("Добавлена заявка: %s", request_data)
            return True
        return False

    def rowCount(self, parent=QModelIndex()):
        return len(self._requests)

    def columnCount(self, parent=QModelIndex()):
        return len(self.HEADERS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        request = self._requests[index.row()]
        if role == Qt.DisplayRole:
            return [
                request["id"],
                request["type"],
                request["status"],
                request["student_fio"],
                request["room_id"],
                request["created_at"]  # Добавляем дату создания
            ][index.column()]
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.HEADERS[section]
        return None

    def update_request_status(self, request_id, new_status):
        """Обновление статуса заявки с проверкой подключения и транзакцией"""
        try:
            # Проверяем, что новый статус допустим
            if new_status not in self.STATUS_VALUES:
                logging.error(f"Попытка установить недопустимый статус: {new_status}")
                return False

            # Проверяем подключение
            if not db.connection or db.connection.closed:
                db.connect()

            # Проверяем существование заявки
            check_query = "SELECT id, type FROM requests WHERE id = %s"
            check_result = db.execute_query(check_query, (request_id,), fetch=True)

            if not check_result or not check_result[0]:
                logging.error(f"Заявка с ID {request_id} не найдена")
                return False

            # Обновляем статус
            update_query = """
            UPDATE requests 
            SET status = %s 
            WHERE id = %s 
            RETURNING id, status
            """
            result = db.execute_query(
                update_query, (new_status, request_id), fetch=True
            )

            if result and result[0]:
                # Обновляем модель
                for i, req in enumerate(self._requests):
                    if req["id"] == str(request_id):
                        self._requests[i]["status"] = new_status
                        # Используем beginResetModel/endResetModel для полного обновления
                        self.beginResetModel()
                        self.endResetModel()
                        break

                # Обновляем комнату и связь студента если заявка Выполнена
                if new_status == "Выполнена":
                    self._update_room_availability(request_id)

                return True
            return False

        except Exception as e:
            logging.error(f"Ошибка при обновлении статуса заявки: {e}")
            return False

    def _update_room_availability(self, request_id):
        """Обновляет доступность комнаты и связывает студента с комнатой при одобрении заявки"""
        for req in self._requests:
            if req["id"] == str(request_id):
                room_id = req["room_id"]
                student_id = None

                # Находим student_id для этой заявки
                query = "SELECT student_id FROM requests WHERE id = %s"
                result = db.execute_query(query, (request_id,), fetch=True)
                if result and result[0]:
                    student_id = result[0][0][0]

                change = 0  # Изменение количества свободных мест

                if req["type"] == "Заселение" and req["status"] == "Выполнена":
                    change = -1
                    # Обновляем room_id студента
                    if student_id:
                        query = "UPDATE students SET room_id = %s WHERE id = %s"
                        db.execute_query(query, (room_id, student_id))
                elif req["type"] == "Выселение" and req["status"] == "Выполнена":
                    change = 1
                    # Получаем текущую комнату студента
                    query = "SELECT room_id FROM students WHERE id = %s"
                    result = db.execute_query(query, (student_id,), fetch=True)
                    if result and result[0] and result[0][0][0]:
                        room_id = result[0][0][0]  # Используем текущую комнату студента
                        # Удаляем связь студента с комнатой (устанавливаем room_id в NULL)
                        query = "UPDATE students SET room_id = NULL WHERE id = %s"
                        db.execute_query(query, (student_id,))

                # Обновляем количество свободных мест в комнате
                if change != 0:
                    query = "UPDATE rooms SET svobodno = svobodno + %s WHERE id = %s"
                    if db.execute_query(query, (change, room_id)):
                        # Находим родительские виджеты для обновления
                        parent = self.parent()
                        while parent and not hasattr(parent, "rooms_tab"):
                            parent = parent.parent()

                        if parent and hasattr(parent, "rooms_tab"):
                            # Обновляем модели комнат и студентов
                            parent.rooms_tab.rooms_model.load_data()
                            parent.students_tab.students_model.load_data()
                break


# Заменяем класс StudentSelectionDialog на следующий:
class StudentSelectionDialog(QDialog):
    def __init__(self, student_model, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор студента")
        self.setMinimumSize(600, 400)

        layout = QVBoxLayout(self)

        # Поле поиска
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Поиск по ФИО или ID...")
        self.search_edit.textChanged.connect(self.filter_students)
        layout.addWidget(self.search_edit)

        # Таблица студентов
        self.table = QTableView()

        # Прокси-модель для фильтрации
        self.proxy_model = (
            StudentsProxyModel()
        )  # Используем уже существующий класс StudentsProxyModel
        self.proxy_model.setSourceModel(student_model)

        self.table.setModel(self.proxy_model)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def filter_students(self, text):
        self.proxy_model.setFilterText(text)

    def get_selected_student(self):
        selection = self.table.selectionModel().selectedRows()
        if selection:
            proxy_index = selection[0]
            source_index = self.proxy_model.mapToSource(proxy_index)
            return self.proxy_model.sourceModel().get_student(source_index.row())
        return None


# Прокси-модель для фильтрации заявок
class RequestsProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.filter_text = ""
        self.status_filter = ""

    def setFilterText(self, text):
        self.filter_text = text
        self.invalidateFilter()

    def setStatusFilter(self, status):
        self.status_filter = status if status != "Все статусы" else ""
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        source_model = self.sourceModel()

        # Фильтр по статусу (если задан)
        if self.status_filter:
            status_index = source_model.index(
                source_row, 2, source_parent
            )  # 2 - колонка статуса
            status = source_model.data(status_index, Qt.DisplayRole)
            if status != self.status_filter:
                return False

        # Фильтр по тексту (если задан)
        if self.filter_text.strip():
            for col in range(source_model.columnCount()):
                index = source_model.index(source_row, col, source_parent)
                data = source_model.data(index, Qt.DisplayRole)
                if data is None:
                    continue
                if self.filter_text.lower() in str(data).lower():
                    return True
            return False

        return True


# Прокси-модель для фильтрации студентов
class StudentsProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.filter_text = ""

    def setFilterText(self, text):
        self.filter_text = text
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        if not self.filter_text.strip():
            return True
        source_model = self.sourceModel()
        for col in range(source_model.columnCount()):
            index = source_model.index(source_row, col, source_parent)
            data = source_model.data(index, Qt.DisplayRole)
            if data is None:
                continue
            if self.filter_text.lower() in str(data).lower():
                return True
        return False


# Диалог добавления/редактирования студента
class AddEditStudentDialog(QDialog):
    def __init__(self, parent=None, student_data=None):
        super().__init__(parent)
        self.setWindowTitle(
            "Добавить студента" if student_data is None else "Редактировать студента"
        )
        self.setModal(True)
        self._is_valid = False

        layout = QFormLayout(self)
        self.fio_edit = QLineEdit()
        self.fio_edit.setPlaceholderText("Введите ФИО (Иванов Иван Иванович)")
        self.fio_edit.textChanged.connect(self.validate_inputs)

        # Добавляем поле для телефона
        self.phone_edit = QLineEdit()
        self.phone_edit.setPlaceholderText("Введите номер телефона")
        self.phone_edit.textChanged.connect(self.validate_inputs)
        
        # Добавляем валидатор для телефона
        phone_validator = QRegularExpressionValidator(
            QRegularExpression(r"^(\+?\d[\d\s\-()]{6,15})?$"), self  # Разрешаем пустую строку
        )
        self.phone_edit.setValidator(phone_validator)

        # Добавляем поле в форму
        layout.addRow("Телефон:", self.phone_edit)

        self.pol_combo = QComboBox()
        self.pol_combo.addItems(["М", "Ж"])

        self.vozrast_spin = QSpinBox(self)
        self.vozrast_spin.setRange(16, 100)
        self.vozrast_spin.setValue(18)
        self.vozrast_spin.valueChanged.connect(self.validate_inputs)

        self.kurs_spin = QSpinBox(self)
        self.kurs_spin.setRange(1, 6)
        self.kurs_spin.setValue(1)
        self.kurs_spin.valueChanged.connect(self.validate_inputs)

        self.fakultet_edit = QLineEdit()
        self.fakultet_edit.setPlaceholderText("Введите факультет")
        self.fakultet_edit.textChanged.connect(self.validate_inputs)

        layout.addRow("ФИО:", self.fio_edit)
        layout.addRow("Пол:", self.pol_combo)
        layout.addRow("Возраст:", self.vozrast_spin)
        layout.addRow("Курс:", self.kurs_spin)
        layout.addRow("Факультет:", self.fakultet_edit)

        button_layout = QHBoxLayout()
        self.ok_button = AnimatedButton("ОК")
        self.ok_button.clicked.connect(self.on_ok)
        cancel_button = QPushButton("Отмена")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(cancel_button)
        layout.addRow(button_layout)

        fio_validator = QRegularExpressionValidator(
            QRegularExpression(r"^[А-ЯЁ][а-яё]+\s[А-ЯЁ][а-яё]+\s[А-ЯЁ][а-яё]+$"), self
        )
        self.fio_edit.setValidator(fio_validator)

        if student_data:
            self.fio_edit.setText(student_data.get("fio", ""))
            index = self.pol_combo.findText(student_data.get("pol", "М"))
            if index >= 0:
                self.pol_combo.setCurrentIndex(index)
            self.vozrast_spin.setValue(int(student_data.get("vozrast", 18)))
            self.kurs_spin.setValue(int(student_data.get("kurs", 1)))
            self.fakultet_edit.setText(student_data.get("fakultet", ""))
            self.phone_edit.setText(student_data.get("number_phone", ""))
        self.validate_inputs()

    def validate_inputs(self):
        valid = True
        if not self.fio_edit.text().strip():
            self.fio_edit.setStyleSheet("border: 1px solid red;")
            valid = False
        else:
            self.fio_edit.setStyleSheet("")

        if not self.fakultet_edit.text().strip():
            self.fakultet_edit.setStyleSheet("border: 1px solid red;")
            valid = False
        else:
            self.fakultet_edit.setStyleSheet("")

        self._is_valid = valid
        self.ok_button.setEnabled(valid)

    def on_ok(self):
        if self._is_valid:
            self.accept()
        else:
            QMessageBox.warning(
                self, "Ошибка", "Заполните обязательные поля корректно."
            )

    def get_student_data(self):
        return {
            "fio": self.fio_edit.text().strip(),
            "pol": self.pol_combo.currentText(),
            "vozrast": self.vozrast_spin.value(),
            "kurs": self.kurs_spin.value(),
            "fakultet": self.fakultet_edit.text().strip(),
            "number_phone": self.phone_edit.text().strip(),  # Добавлено поле телефона
        }


# Кастомная кнопка с анимацией
class AnimatedButton(QPushButton):
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.effect = QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(self.effect)
        self.animation = QPropertyAnimation(self.effect, b"opacity")
        self.animation.setDuration(150)
        self.animation.setStartValue(0.7)
        self.animation.setEndValue(1.0)
        self.animation.setEasingCurve(QEasingCurve.InOutQuad)
        self.clicked.connect(self.start_animation)

    def start_animation(self):
        self.animation.start()


# Диалог создания заявки на заселение
class ZaselRequestDialog(QDialog):
    def __init__(self, student_model, room_model, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Создать заявку на заселение")
        self.student_model = student_model
        self.room_model = room_model
        layout = QFormLayout(self)

        # Кнопка выбора студента
        self.select_student_btn = QPushButton("Выбрать студента...")
        self.select_student_btn.clicked.connect(self.select_student)
        self.selected_student = None
        self.student_label = QLabel("Не выбран")
        layout.addRow("Студент:", self.select_student_btn)
        layout.addRow("Выбран:", self.student_label)

        self.room_combo = QComboBox(self)
        for i in range(room_model.rowCount()):
            room = room_model._rooms[i]
            try:
                if int(room.get("svobodno", "0")) > 0:
                    display = f"{room.get('id')} - Этаж: {room.get('etazh')}, Мест: {room.get('kol_mest')}, Свободно: {room.get('svobodno')}"
                    self.room_combo.addItem(display, room)
            except Exception:
                continue

        layout.addRow("Комната:", self.room_combo)

        button_layout = QHBoxLayout()
        ok_button = AnimatedButton("ОК")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Отмена")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        layout.addRow(button_layout)

    def select_student(self):
        dialog = StudentSelectionDialog(self.student_model, self)
        if dialog.exec() == QDialog.Accepted:
            self.selected_student = dialog.get_selected_student()
            if self.selected_student:
                self.student_label.setText(
                    f"{self.selected_student['id']} - {self.selected_student['fio']} (Комната: {self.selected_student['room_id']})"
                )

    def get_request_data(self):
        if not self.selected_student:
            return None

        room = self.room_combo.currentData()
        return {
            "type": "Заселение",
            "status": "Создана",
            "student_id": self.selected_student["id"],
            "student_fio": self.selected_student["fio"],
            "room_id": room["id"],
        }


# Диалог создания заявки на выселение
class VyselRequestDialog(QDialog):
    def __init__(self, student_model, room_model, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Создать заявку на выселение")
        self.student_model = student_model
        self.room_model = room_model
        layout = QFormLayout(self)

        # Кнопка выбора студента
        self.select_student_btn = QPushButton("Выбрать студента...")
        self.select_student_btn.clicked.connect(self.select_student)
        self.selected_student = None
        self.student_label = QLabel("Не выбран")
        layout.addRow("Студент:", self.select_student_btn)
        layout.addRow("Выбран:", self.student_label)

        # Комната будет определяться автоматически из данных студента
        self.room_label = QLabel("Будет определена автоматически")
        layout.addRow("Комната:", self.room_label)

        button_layout = QHBoxLayout()
        ok_button = AnimatedButton("ОК")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Отмена")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        layout.addRow(button_layout)

    def select_student(self):
        dialog = StudentSelectionDialog(self.student_model, self)
        if dialog.exec() == QDialog.Accepted:
            self.selected_student = dialog.get_selected_student()
            if self.selected_student:
                room_info = self.selected_student["room_id"]
                self.student_label.setText(
                    f"{self.selected_student['id']} - {self.selected_student['fio']} (Комната: {room_info})"
                )
                # Получаем текущую комнату студента
                query = "SELECT room_id FROM students WHERE id = %s"
                result = db.execute_query(
                    query, (int(self.selected_student["id"]),), fetch=True
                )
                if result and result[0] and result[0][0][0]:
                    room_id = result[0][0][0]
                    self.room_label.setText(str(room_id))
                else:
                    self.room_label.setText("Студент не заселен")

    def get_request_data(self):
        if not self.selected_student:
            return None

        # Получаем текущую комнату студента
        query = "SELECT room_id FROM students WHERE id = %s"
        result = db.execute_query(
            query, (int(self.selected_student["id"]),), fetch=True
        )
        if not result or not result[0] or not result[0][0][0]:
            QMessageBox.warning(self, "Ошибка", "Студент не заселен в комнату")
            return None  # Просто возвращаем None без вылета

        room_id = result[0][0][0]
        return {
            "type": "Выселение",
            "status": "Создана",
            "student_id": self.selected_student["id"],
            "student_fio": self.selected_student["fio"],
            "room_id": room_id,
        }


# Виджет для отображения списка студентов
class StudentsWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)

        filter_layout = QHBoxLayout()
        filter_label = QLabel("Фильтр:")
        self.filter_line_edit = QLineEdit()
        self.filter_line_edit.setPlaceholderText("Введите текст для поиска")
        self.filter_line_edit.textChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.filter_line_edit, 1)
        layout.addLayout(filter_layout)

        # Добавляем кнопку обновления в верхней части
        refresh_layout = QHBoxLayout()
        self.refresh_button = AnimatedButton("Обновить")
        self.refresh_button.clicked.connect(self.refresh_data)
        refresh_layout.addWidget(self.refresh_button)
        refresh_layout.addStretch()
        layout.addLayout(refresh_layout)

        filter_layout = QHBoxLayout()

        self.students_model = StudentsTableModel()
        self.proxy_model = StudentsProxyModel()
        self.proxy_model.setSourceModel(self.students_model)

        self.table_view = QTableView()
        self.table_view.setModel(self.proxy_model)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        layout.addWidget(self.table_view)

        button_layout = QHBoxLayout()
        add_button = AnimatedButton("Добавить студента")
        add_button.clicked.connect(self.add_student)
        edit_button = AnimatedButton("Редактировать")
        edit_button.clicked.connect(self.edit_student)
        edit_multiple_button = AnimatedButton("Групповое редактирование")
        edit_multiple_button.clicked.connect(self.edit_multiple_students)
        delete_button = AnimatedButton("Удалить")
        delete_button.clicked.connect(self.delete_student)
        delete_multiple_button = AnimatedButton("Групповое удаление")
        delete_multiple_button.clicked.connect(self.delete_multiple_students)
        
        button_layout.addWidget(add_button)
        button_layout.addWidget(edit_button)
        button_layout.addWidget(edit_multiple_button)
        button_layout.addWidget(delete_button)
        button_layout.addWidget(delete_multiple_button)
        
        layout.addLayout(button_layout)

    def on_filter_changed(self, text):
        self.proxy_model.setFilterText(text)
        logging.info("Применён фильтр: %s", text)

    def refresh_data(self):
        """Обновляет данные студентов"""
        self.students_model.load_data()

    def add_student(self):
        dialog = AddEditStudentDialog(self)
        if dialog.exec() == QDialog.Accepted:
            student = dialog.get_student_data()
            if self.students_model.add_student(student):
                # Принудительно обновляем представление
                self.students_model.layoutChanged.emit()
                QMessageBox.information(self, "Успех", "Студент успешно добавлен.")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось добавить студента.")

    def edit_student(self):
        selection = self.table_view.selectionModel().selectedRows()
        if not selection:
            QMessageBox.warning(self, "Ошибка", "Выберите студента для редактирования.")
            return

        proxy_index = selection[0]
        source_index = self.proxy_model.mapToSource(proxy_index)
        row = source_index.row()
        student = self.students_model.get_student(row)

        if not student:
            QMessageBox.warning(self, "Ошибка", "Не удалось получить данные студента.")
            return

        dialog = AddEditStudentDialog(self, student)
        if dialog.exec() == QDialog.Accepted:
            updated_student = dialog.get_student_data()
            updated_student["id"] = student["id"]
            if self.students_model.update_student(row, updated_student):
                QMessageBox.information(self, "Успех", "Данные студента обновлены.")
            else:
                QMessageBox.warning(
                    self, "Ошибка", "Не удалось обновить данные студента."
                )

    def delete_student(self):
        selection = self.table_view.selectionModel().selectedRows()
        if not selection:
            QMessageBox.warning(self, "Ошибка", "Выберите студента для удаления.")
            return

        proxy_index = selection[0]
        source_index = self.proxy_model.mapToSource(proxy_index)
        row = source_index.row()

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить студента?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            if self.students_model.remove_student(row):
                QMessageBox.information(self, "Успех", "Студент удалён.")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить студента.")

    def edit_multiple_students(self):
        """Редактирование нескольких выбранных студентов"""
        selection = self.table_view.selectionModel().selectedRows()
        if len(selection) < 2:
            QMessageBox.warning(self, "Ошибка", "Выберите несколько студентов для группового редактирования.")
            return
            
        # Получаем список выбранных строк
        rows = [self.proxy_model.mapToSource(index).row() for index in selection]
        
        # Создаем диалог для выбора полей для редактирования
        dialog = QDialog(self)
        dialog.setWindowTitle("Групповое редактирование")
        layout = QFormLayout(dialog)
        
        # Поля для редактирования
        fields = {
            "fio": ("ФИО", QLineEdit()),
            "pol": ("Пол", QComboBox()),
            "vozrast": ("Возраст", QSpinBox()),
            "kurs": ("Курс", QSpinBox()),
            "fakultet": ("Факультет", QLineEdit()),
            "number_phone": ("Телефон", QLineEdit())
        }
        
        # Настройка виджетов
        fields["pol"][1].addItems(["М", "Ж"])
        fields["vozrast"][1].setRange(16, 100)
        fields["kurs"][1].setRange(1, 6)
        
        # Добавляем чекбоксы для выбора полей
        checkboxes = {}
        for key, (label, widget) in fields.items():
            hbox = QHBoxLayout()
            checkbox = QCheckBox("Изменить")
            checkboxes[key] = checkbox
            hbox.addWidget(checkbox)
            hbox.addWidget(QLabel(label))
            hbox.addWidget(widget)
            layout.addRow(hbox)
        
        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            # Собираем данные для обновления
            update_data = {}
            for key in fields:
                if checkboxes[key].isChecked():
                    widget = fields[key][1]
                    if isinstance(widget, QLineEdit):
                        update_data[key] = widget.text().strip()
                    elif isinstance(widget, QComboBox):
                        update_data[key] = widget.currentText()
                    elif isinstance(widget, QSpinBox):
                        update_data[key] = widget.value()
            
            if update_data:
                if self.students_model.update_multiple_students(rows, update_data):
                    QMessageBox.information(
                        self, 
                        "Успех", 
                        f"Успешно обновлено {len(rows)} студентов."
                    )
                else:
                    QMessageBox.warning(
                        self, 
                        "Ошибка", 
                        "Не удалось обновить студентов."
                    )
    
    def delete_multiple_students(self):
        """Удаление нескольких выбранных студентов"""
        selection = self.table_view.selectionModel().selectedRows()
        if len(selection) < 2:
            QMessageBox.warning(self, "Ошибка", "Выберите несколько студентов для группового удаления.")
            return
            
        rows = [self.proxy_model.mapToSource(index).row() for index in selection]
        
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить {len(rows)} выбранных студентов?",
            QMessageBox.Yes | QMessageBox.No,
        )
        
        if reply == QMessageBox.Yes:
            if self.students_model.remove_multiple_students(rows):
                QMessageBox.information(
                    self, 
                    "Успех", 
                    f"Удалено {len(rows)} студентов."
                )
            else:
                QMessageBox.warning(
                    self, 
                    "Ошибка", 
                    "Не удалось удалить студентов."
                )


# Виджет для отображения списка комнат
# Виджет для отображения списка комнат
class RoomsWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)

        # Добавляем строку поиска
        filter_layout = QHBoxLayout()
        filter_label = QLabel("Поиск:")
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText("Введите ID комнаты или этаж...")
        self.filter_edit.textChanged.connect(self.apply_filter)
        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.filter_edit)
        layout.addLayout(filter_layout)

        # Кнопка обновления
        self.refresh_button = AnimatedButton("Обновить список")
        self.refresh_button.clicked.connect(self.refresh_data)
        layout.addWidget(self.refresh_button)

        # Таблица комнат
        self.table = QTableView()
        self.table.setStyleSheet(
            """
            QTableView {
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background-color: white;
            }
            QHeaderView::section {
                background-color: #ecf0f1; 
                color: #2c3e50;            
                padding: 8px;
            }
        """
        )
        self.rooms_model = RoomsTableModel()
        
        # Прокси-модель для фильтрации
        self.proxy_model = QSortFilterProxyModel()
        self.proxy_model.setSourceModel(self.rooms_model)
        self.proxy_model.setFilterCaseSensitivity(Qt.CaseInsensitive)
        # Фильтрация будет выполняться вручную через filterAcceptsRow
        self.proxy_model.setFilterKeyColumn(-1)  # -1 означает, что фильтрация будет выполняться для всех столбцов
        
        self.table.setModel(self.proxy_model)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setSelectionMode(QTableView.SingleSelection)
        self.table.setMinimumHeight(200)

        # Подключаем обработчик выбора
        self.table.selectionModel().selectionChanged.connect(self.show_room_students)
        layout.addWidget(self.table)

        # Остальной код без изменений...
        # Виджет для отображения студентов
        self.students_frame = QFrame()
        self.students_frame.setStyleSheet(
            """
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 15px;
            }
        """
        )
        students_layout = QVBoxLayout(self.students_frame)
        students_layout.setContentsMargins(10, 10, 10, 10)

        self.students_title = QLabel("Выберите комнату")
        self.students_title.setStyleSheet(
            """
            font-size: 14px;
            font-weight: bold;
            color: #2c3e50;
            padding-bottom: 10px;
        """
        )
        students_layout.addWidget(self.students_title)

        self.students_list = QLabel()
        self.students_list.setStyleSheet(
            """
            font-size: 13px;
            color: #495057;
        """
        )
        self.students_list.setWordWrap(True)
        students_layout.addWidget(self.students_list)

        layout.addWidget(self.students_frame)

    def apply_filter(self, text):
        """Применяет фильтр к данным комнат (только по ID или этажу)"""
        self.proxy_model.setFilterFixedString(text)
        logging.info(f"Применён фильтр комнат: {text}")

    def filterAcceptsRow(self, source_row, source_parent):
        """Переопределяем метод фильтрации для поиска только по ID и этажу"""
        if not self.filter_edit.text():
            return True
            
        text = self.filter_edit.text().lower()
        model = self.proxy_model.sourceModel()
        
        # Проверяем ID комнаты (первый столбец)
        id_index = model.index(source_row, 0, source_parent)
        id_data = model.data(id_index, Qt.DisplayRole)
        if text in str(id_data).lower():
            return True
            
        # Проверяем этаж (второй столбец)
        floor_index = model.index(source_row, 1, source_parent)
        floor_data = model.data(floor_index, Qt.DisplayRole)
        if text in str(floor_data).lower():
            return True
            
        return False

    def show_room_students(self):
        """Отображает список студентов в выбранной комнате"""
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            self.students_title.setText("Выберите комнату")
            self.students_list.setText(
                "Для просмотра списка проживающих выберите комнату из таблицы выше"
            )
            return

        # Получаем индекс из прокси-модели и преобразуем его в индекс исходной модели
        proxy_index = selected[0]
        source_index = self.proxy_model.mapToSource(proxy_index)
        
        room_id = self.rooms_model.index(source_index.row(), 0).data()

        # Получаем данные о комнате
        room_query = "SELECT etazh, kol_mest, svobodno FROM rooms WHERE id = %s"
        room_data = db.execute_query(room_query, (room_id,), fetch=True)

        # Получаем список студентов
        students_query = """
        SELECT fio, kurs, fakultet 
        FROM students 
        WHERE room_id = %s
        ORDER BY fio
        """
        students_data = db.execute_query(students_query, (room_id,), fetch=True)

        # Формируем текст
        if room_data and room_data[0]:
            etazh, kol_mest, svobodno = room_data[0][0]
            occupied = kol_mest - svobodno

            room_info = f"<b>Комната {room_id}</b> (Этаж {etazh})<br>"
            room_info += f"Мест: {kol_mest}, свободно: {svobodno}<br><br>"

            if students_data and students_data[0]:
                students = students_data[0]
                students_list = "<b>Проживают:</b><br>"
                students_list += "<br>".join(
                    [f"• {s[0]} ({s[2]}, {s[1]} курс)" for s in students]
                )
                students_list += f"<br><br>Всего проживает: {occupied}"
            else:
                students_list = "В комнате никто не проживает"

            self.students_title.setText(f"Информация о комнате {room_id}")
            self.students_list.setText(room_info + students_list)
        else:
            self.students_title.setText("Ошибка")
            self.students_list.setText("Не удалось загрузить информацию о комнате")

    def refresh_data(self):
        """Обновляет данные в таблице комнат"""
        self.rooms_model.load_data()
        self.students_title.setText("Выберите комнату")
        self.students_list.setText(
            "Данные обновлены. Выберите комнату для просмотра информации."
        )
        QMessageBox.information(self, "Обновление", "Данные о комнатах обновлены.")

# Виджет для работы с заявками
class RequestsWidget(QWidget):
    def __init__(self, student_model, room_model):
        super().__init__()
        self.student_model = student_model
        self.room_model = room_model

        layout = QVBoxLayout(self)

        # Добавляем фильтры
        filter_layout = QHBoxLayout()

        # Фильтр по тексту
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText(
            "Поиск по ID, типу, студенту или комнате..."
        )
        self.filter_edit.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Поиск:"))
        filter_layout.addWidget(self.filter_edit)

        # Фильтр по статусу
        self.status_combo = QComboBox()
        self.status_combo.addItem("Все статусы")  # Первый элемент - "Все статусы"
        self.status_combo.addItems(RequestsTableModel.STATUS_VALUES)
        self.status_combo.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Статус:"))
        filter_layout.addWidget(self.status_combo)

        layout.addLayout(filter_layout)

        # Кнопки создания заявок
        create_buttons = QHBoxLayout()
        add_zasel_btn = AnimatedButton("Создать заявку на заселение")
        add_zasel_btn.clicked.connect(self.create_zasel_request)
        add_vysel_btn = AnimatedButton("Создать заявку на выселение")
        add_vysel_btn.clicked.connect(self.create_vysel_request)
        create_buttons.addWidget(add_zasel_btn)
        create_buttons.addWidget(add_vysel_btn)
        create_buttons.addWidget(AnimatedButton("Обновить"))
        create_buttons.itemAt(create_buttons.count() - 1).widget().clicked.connect(
            self.refresh_data
        )
        layout.addLayout(create_buttons)

        # Кнопки обработки заявок
        process_buttons = QHBoxLayout()
        self.process_btn = AnimatedButton("В обработку")
        self.process_btn.clicked.connect(lambda: self.process_request("В обработке"))
        self.approve_btn = AnimatedButton("Одобрить")
        self.approve_btn.clicked.connect(lambda: self.process_request("Выполнена"))
        self.reject_btn = AnimatedButton("Отклонить")
        self.reject_btn.clicked.connect(lambda: self.process_request("Отклонена"))
        process_buttons.addWidget(self.process_btn)
        process_buttons.addWidget(self.approve_btn)
        process_buttons.addWidget(self.reject_btn)
        layout.addLayout(process_buttons)

        # Таблица заявок с прокси-моделью
        self.requests_table = QTableView()
        self.requests_model = RequestsTableModel()
        self.proxy_model = RequestsProxyModel()
        self.proxy_model.setSourceModel(self.requests_model)
        self.requests_table.setModel(self.proxy_model)
        self.requests_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.requests_table.setSelectionBehavior(QTableView.SelectRows)
        layout.addWidget(self.requests_table)

        # Обновляем кнопки при выборе заявки
        self.requests_table.selectionModel().selectionChanged.connect(
            self.update_buttons_state
        )
        self.update_buttons_state()

        # Кнопка теста подключения (остается без изменений)
        test_btn = QPushButton("Проверить подключение к БД")
        test_btn.clicked.connect(self.test_db_connection)
        layout.addWidget(test_btn)

    def apply_filters(self):
        """Применяет фильтры к прокси-модели"""
        self.proxy_model.setFilterText(self.filter_edit.text())
        self.proxy_model.setStatusFilter(self.status_combo.currentText())

    def refresh_data(self):
        """Обновляет данные заявок с сохранением сортировки"""
        self.requests_model.load_data()
        # Прокручиваем таблицу вверх после обновления
        self.requests_table.scrollToTop()

    def update_buttons_state(self):
        """Обновляет состояние кнопок в зависимости от выбора и текущего статуса"""
        selected = self.requests_table.selectionModel().selectedRows()
        has_selection = bool(selected)

        if has_selection:
            # Получаем индекс из прокси-модели и преобразуем его в индекс исходной модели
            proxy_index = selected[0]
            source_index = self.proxy_model.mapToSource(proxy_index)

            # Получаем статус заявки из исходной модели
            status_index = self.requests_model.index(
                source_index.row(), 2
            )  # 2 - колонка статуса
            status = self.requests_model.data(status_index, Qt.DisplayRole)

            # Настройка доступности кнопок в зависимости от текущего статуса
            self.process_btn.setEnabled(status == "Создана")
            self.approve_btn.setEnabled(status in ["Создана", "В обработке"])
            self.reject_btn.setEnabled(status in ["Создана", "В обработке"])
        else:
            self.process_btn.setEnabled(False)
            self.approve_btn.setEnabled(False)
            self.reject_btn.setEnabled(False)

    def process_request(self, new_status):
        """Обработка заявки (изменение статуса)"""
        selected = self.requests_table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите заявку для обработки")
            return

        # Получаем индекс из прокси-модели и преобразуем его в индекс исходной модели
        proxy_index = selected[0]
        source_index = self.proxy_model.mapToSource(proxy_index)

        request_id = self.requests_model.index(source_index.row(), 0).data()
        request_type = self.requests_model.index(source_index.row(), 1).data()
        current_status = self.requests_model.index(source_index.row(), 2).data()

        # Подтверждение действия
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите изменить статус заявки {request_id} ({request_type}) с '{current_status}' на '{new_status}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                if self.requests_model.update_request_status(request_id, new_status):
                    QMessageBox.information(
                        self,
                        "Успех",
                        f"Статус заявки {request_id} изменен на '{new_status}'",
                    )
                    # Принудительно обновляем фильтр
                    self.apply_filters()
                else:
                    QMessageBox.warning(
                        self,
                        "Ошибка",
                        "Не удалось обновить статус заявки. Проверьте логи.",
                    )
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def test_db_connection(self):
        """Тестирование подключения к базе данных"""
        try:
            result = db.execute_query("SELECT 1", fetch=True)
            if result:
                QMessageBox.information(
                    self,
                    "Проверка подключения",
                    "Подключение к базе данных работает нормально",
                )
            else:
                QMessageBox.warning(
                    self, "Проверка подключения", "Не удалось выполнить тестовый запрос"
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка подключения",
                f"Ошибка при подключении к базе данных: {str(e)}",
            )

    def create_zasel_request(self):
        dialog = ZaselRequestDialog(self.student_model, self.room_model, self)
        if dialog.exec() == QDialog.Accepted:
            request = dialog.get_request_data()
            if self.requests_model.add_request(request):
                QMessageBox.information(
                    self, "Заселение", "Заявка на заселение создана."
                )
            else:
                QMessageBox.warning(
                    self, "Ошибка", "Не удалось создать заявку на заселение."
                )

    def create_vysel_request(self):
        dialog = VyselRequestDialog(self.student_model, self.room_model, self)
        if dialog.exec() == QDialog.Accepted:
            request = dialog.get_request_data()
            if request:  # Проверяем, что request не None
                if self.requests_model.add_request(request):
                    QMessageBox.information(
                        self, "Выселение", "Заявка на выселение создана."
                    )
                else:
                    QMessageBox.warning(
                        self, "Ошибка", "Не удалось создать заявку на выселение."
                    )


# Асинхронное задание для генерации отчёта
class WorkerSignals(QObject):
    progress = Signal(int)
    finished = Signal(bool)
    error = Signal(str)


class ReportWorkerRunnable(QRunnable):
    def __init__(self):
        super().__init__()
        self.signals = WorkerSignals()
        self._is_cancelled = False

    def run(self):
        try:
            # Шаг 1: Получение данных
            self.signals.progress.emit(10)
            if self._is_cancelled:
                self.signals.finished.emit(False)
                return

            # Получаем данные студентов с информацией о последней заявке
            query = """
            SELECT 
                s.id AS student_id,
                s.fio,
                s.pol,
                s.vozrast,
                s.kurs,
                s.fakultet,
                COALESCE(r.id::text, 'не заселен') AS room_number,
                COALESCE(
                    (SELECT CONCAT(req.type, ' - ', req.status)
                    FROM requests req 
                    WHERE req.student_id = s.id
                    ORDER BY req.created_at DESC
                    LIMIT 1),
                    'нет активных заявок'
                ) AS last_request_info
            FROM students s
            LEFT JOIN rooms r ON s.room_id = r.id
            ORDER BY s.fio
            """

            result = db.execute_query(query, fetch=True)
            if not result or not isinstance(result, tuple) or len(result) != 2:
                self.signals.error.emit("Ошибка при получении данных из БД")
                self.signals.finished.emit(False)
                return

            students_data, columns = result

            self.signals.progress.emit(50)
            if self._is_cancelled:
                self.signals.finished.emit(False)
                return

            # Шаг 2: Создание Excel-файла
            wb = Workbook()
            ws = wb.active
            ws.title = "Студенты"

            # Заголовки столбцов
            headers = [
                "ID",
                "ФИО",
                "Пол",
                "Возраст",
                "Курс",
                "Факультет",
                "Комната",
                "Последняя заявка",
            ]

            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(
                start_color="3498db", end_color="3498db", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Заголовок отчета
            ws.merge_cells("A1:H1")
            ws["A1"] = "Отчет по студентам общежития"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")

            # Заголовки столбцов
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = max(len(header) + 2, 12)

            # Данные студентов
            if students_data:
                for row_num, row_data in enumerate(students_data, 3):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.border = thin_border

                        column_letter = get_column_letter(col_num)
                        current_width = ws.column_dimensions[column_letter].width
                        cell_length = len(str(cell_value)) + 2
                        if cell_length > current_width:
                            ws.column_dimensions[column_letter].width = cell_length
            else:
                ws["A3"] = "Нет данных о студентах"

            # Замораживаем заголовки
            ws.freeze_panes = "A3"

            # Добавляем дату создания отчета
            last_row = len(students_data) + 3 if students_data else 4
            ws["A" + str(last_row)] = "Отчет создан:"
            ws["B" + str(last_row)] = datetime.now().strftime("%d.%m.%Y %H:%M")

            # Сохранение файла
            filename = (
                f"students_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            wb.save(filename)

            self.signals.progress.emit(100)
            self.signals.finished.emit(True)

        except Exception as e:
            logging.error(f"Ошибка при генерации отчёта: {e}")
            self.signals.error.emit(f"Ошибка: {str(e)}")
            self.signals.finished.emit(False)


# Виджет для отчётности
class ReportsWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Добавляем строку с кнопкой обновления
        refresh_layout = QHBoxLayout()
        self.refresh_button = AnimatedButton("Обновить статистику")
        self.refresh_button.clicked.connect(self.load_statistics)
        refresh_layout.addWidget(self.refresh_button)
        refresh_layout.addStretch()
        layout.addLayout(refresh_layout)
        
        # Добавляем заголовок
        title = QLabel("Статистика общежития")
        title.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
            }
        """)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Остальной код остается без изменений...
        # Первая строка карточек
        row1 = QHBoxLayout()
        row1.setSpacing(10)
        self.total_students_card = self._create_stat_card("Всего студентов", "Загрузка...")
        self.total_rooms_card = self._create_stat_card("Всего комнат", "Загрузка...")
        self.occupied_rooms_card = self._create_stat_card("Занято комнат", "Загрузка...")
        self.free_rooms_card = self._create_stat_card("Свободно мест", "Загрузка...")
        row1.addWidget(self.total_students_card)
        row1.addWidget(self.total_rooms_card)
        row1.addWidget(self.occupied_rooms_card)
        row1.addWidget(self.free_rooms_card)
        layout.addLayout(row1)
        
        # Вторая строка карточек
        row2 = QHBoxLayout()
        row2.setSpacing(10)
        self.avg_age_card = self._create_stat_card("Средний возраст", "Загрузка...")
        self.by_gender_card = self._create_stat_card("По полу", "Загрузка...")
        self.by_course_card = self._create_stat_card("По курсам", "Загрузка...")
        self.by_faculty_card = self._create_stat_card("По факультетам", "Загрузка...")
        row2.addWidget(self.avg_age_card)
        row2.addWidget(self.by_gender_card)
        row2.addWidget(self.by_course_card)
        row2.addWidget(self.by_faculty_card)
        layout.addLayout(row2)
        
        # Кнопка экспорта
        self.export_button = AnimatedButton("Выгрузить полный отчёт в Excel")
        self.export_button.clicked.connect(self.export_report)
        layout.addWidget(self.export_button, alignment=Qt.AlignCenter)
        
        self.progress_dialog = None
        
        # Загружаем статистику сразу
        self.load_statistics()
        
    def _create_stat_card(self, title, value):
        """Создает карточку для отображения статистики"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border: 1px solid #d6d6d6;
                border-radius: 8px;
                padding: 10px;
                min-width: 180px;
                min-height: 80px;
            }
        """)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(5)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #3498db;
                font-size: 14px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        
        value_label = QLabel(value)
        value_label.setObjectName("value_label")
        value_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                color: #2c3e50;
            }
        """)
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setWordWrap(True)
        
        layout.addWidget(title_label)
        layout.addWidget(value_label, stretch=1)
        
        return card

    def load_statistics(self):
        """Загружает статистические данные из базы"""
        try:
            # 1. Общее количество студентов
            query = "SELECT COUNT(*) FROM students"
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                count = result[0][0][0]
                self._update_card(self.total_students_card, str(count))
            
            # 2. Общее количество комнат
            query = "SELECT COUNT(*) FROM rooms"
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                count = result[0][0][0]
                self._update_card(self.total_rooms_card, str(count))
            
            # 3. Количество занятых комнат
            query = """
            SELECT COUNT(DISTINCT room_id) 
            FROM students 
            WHERE room_id IS NOT NULL
            """
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                count = result[0][0][0]
                self._update_card(self.occupied_rooms_card, str(count))
            
            # 4. Количество свободных мест
            query = "SELECT SUM(svobodno) FROM rooms"
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                count = result[0][0][0] or 0
                self._update_card(self.free_rooms_card, str(count))
            
            # 5. Средний возраст студентов
            query = "SELECT ROUND(AVG(vozrast), 1) FROM students"
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                avg = result[0][0][0]
                self._update_card(self.avg_age_card, f"{avg} лет")
            
            # 6. Распределение по полу
            query = """
            SELECT pol, COUNT(*) 
            FROM students 
            GROUP BY pol 
            ORDER BY pol
            """
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                gender_stats = {row[0]: row[1] for row in result[0]}
                text = f"Мужчин: {gender_stats.get('М', 0)}\nЖенщин: {gender_stats.get('Ж', 0)}"
                self._update_card(self.by_gender_card, text)
            
            # 7. Распределение по курсам
            query = """
            SELECT kurs, COUNT(*) 
            FROM students 
            GROUP BY kurs 
            ORDER BY kurs
            """
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                text = "\n".join([f"{row[0]} курс: {row[1]}" for row in result[0]])
                self._update_card(self.by_course_card, text)
            
            # 8. Распределение по факультетам (топ 3)
            query = """
            SELECT fakultet, COUNT(*) as count 
            FROM students 
            GROUP BY fakultet 
            ORDER BY count DESC 
            LIMIT 3
            """
            result = db.execute_query(query, fetch=True)
            if result and result[0]:
                text = "\n".join([f"{row[0]}: {row[1]}" for row in result[0]])
                if len(result[0]) == 3:
                    text += "\n..."
                self._update_card(self.by_faculty_card, text)
                
        except Exception as e:
            logging.error(f"Ошибка при загрузке статистики: {e}")
            for card in [
                self.total_students_card, self.total_rooms_card, 
                self.occupied_rooms_card, self.free_rooms_card,
                self.avg_age_card, self.by_gender_card, 
                self.by_course_card, self.by_faculty_card
            ]:
                self._update_card(card, "Ошибка")

    def _update_card(self, card, value):
        """Обновляет значение в карточке статистики"""
        value_label = card.findChild(QLabel, "value_label")
        if value_label:
            value_label.setText(str(value))

    def export_report(self):
        self.progress_dialog = QProgressDialog(
            "Подготовка данных для выгрузки...", "Отмена", 0, 100, self
        )
        self.progress_dialog.setWindowTitle("Экспорт в Excel")
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setAutoClose(True)

        self.worker = ReportWorkerRunnable()
        self.worker.signals.progress.connect(self.progress_dialog.setValue)
        self.worker.signals.finished.connect(self.report_finished)
        self.worker.signals.error.connect(self.report_error)
        self.progress_dialog.canceled.connect(
            lambda: setattr(self.worker, "_is_cancelled", True)
        )

        QThreadPool.globalInstance().start(self.worker)

    def report_finished(self, success):
        if success:
            QMessageBox.information(
                self, "Успех", "Отчёт успешно выгружен в файл 'report.xlsx'"
            )
        self.progress_dialog.close()

    def report_error(self, error_msg):
        QMessageBox.critical(self, "Ошибка", f"Ошибка при выгрузке отчёта: {error_msg}")
        self.progress_dialog.close()
        
# Главное окно приложения
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Система управления общежитием")
        self.setMinimumSize(900, 700)
        self._create_tabs()
        self._create_menu()
        self.setStyleSheet(APP_QSS)
        # Связываем обновление студентов при изменении заявок
        self.requests_tab.requests_model.dataChanged.connect(self._update_students_view)

    def _update_students_view(self):
        """Обновляет вид студентов при изменении заявок"""
        self.students_tab.students_model.load_data()

    def _create_tabs(self):
        self.tabs = QTabWidget()
        self.students_tab = StudentsWidget()
        self.rooms_tab = RoomsWidget()
        self.requests_tab = RequestsWidget(
            self.students_tab.students_model, self.rooms_tab.rooms_model
        )
        self.reports_tab = ReportsWidget()

        self.tabs.addTab(self.students_tab, "Студенты")
        self.tabs.addTab(self.rooms_tab, "Комнаты")
        self.tabs.addTab(self.requests_tab, "Заявки")
        self.tabs.addTab(self.reports_tab, "Отчётность")

        self.setCentralWidget(self.tabs)

    def _create_menu(self):
        menu_bar = self.menuBar()

        file_menu = menu_bar.addMenu("Файл")
        exit_action = QAction("Выход", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        logs_action = QAction("Просмотр логов", self)
        logs_action.triggered.connect(self.show_logs)
        file_menu.addAction(logs_action)

        help_menu = menu_bar.addMenu("Справка")
        
        # Добавляем новый пункт меню с туториалом
        tutorial_action = QAction("Руководство пользователя", self)
        tutorial_action.triggered.connect(self.show_tutorial)
        help_menu.addAction(tutorial_action)
        
        # Добавляем разделитель
        help_menu.addSeparator()
        
        about_action = QAction("О программе", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    # Добавим новый метод для показа туториала
    def show_tutorial(self):
        tutorial_text = """
        <h2>Руководство пользователя</h2>
        
        <h3>Основные разделы:</h3>
        <ul>
            <li><b>Студенты</b> - управление списком проживающих</li>
            <li><b>Комнаты</b> - просмотр информации о комнатах и их заселенности</li>
            <li><b>Заявки</b> - обработка заявок на заселение/выселение</li>
            <li><b>Отчётность</b> - статистика и выгрузка данных</li>
        </ul>
        
        <h3>Быстрые подсказки:</h3>
        <ol>
            <li>Для добавления студента нажмите кнопку "Добавить студента"</li>
            <li>Чтобы создать заявку, выберите соответствующую кнопку в разделе "Заявки"</li>
            <li>Статус заявки можно изменить кнопками "В обработку", "Одобрить" или "Отклонить"</li>
            <li>Для поиска используйте поле фильтра в каждом разделе</li>
            <li>Обновить данные можно кнопкой "Обновить" в каждом разделе</li>
        </ol>
        
        <h3>Частые вопросы:</h3>
        <p><b>Q:</b> Как заселить студента?</p>
        <p><b>A:</b> 1) Убедитесь, что студент добавлен в систему. 2) В разделе "Заявки" создайте заявку на заселение. 3) Одобрите заявку.</p>
        
        <p><b>Q:</b> Почему не отображаются комнаты?</p>
        <p><b>A:</b> Проверьте подключение к базе данных (кнопка внизу раздела "Заявки").</p>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("Руководство пользователя")
        msg.setTextFormat(Qt.RichText)
        msg.setText(tutorial_text)
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

    def show_about(self):
        about_text = """
        <h2>Система управления общежитием</h2>
        <p>Версия 2.3</p>
        <p>© 2025 Университет</p>
        <p>Для получения помощи воспользуйтесь разделом <b>"Руководство пользователя"</b> в меню Справка.</p>
        <p>При возникновении проблем просмотрите логи (меню Файл → Просмотр логов).</p>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("О программе")
        msg.setTextFormat(Qt.RichText)
        msg.setText(about_text)
        msg.setIcon(QMessageBox.Information)
        msg.exec()
        
    def show_logs(self):
        dlg = LogViewerDialog(self)
        dlg.exec()

    def closeEvent(self, event):
        db.close()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    default_font = QFont("Segoe UI", 10)
    app.setFont(default_font)

    # Проверка подключения к базе данных
    try:
        # Создаем таблицы, если они не существуют
        create_tables_query = """
        CREATE TABLE IF NOT EXISTS students (
            id SERIAL PRIMARY KEY,
            fio VARCHAR(100) NOT NULL,
            pol VARCHAR(1) NOT NULL,
            vozrast INTEGER NOT NULL,
            kurs INTEGER NOT NULL,
            fakultet VARCHAR(50) NOT NULL,
            number_phone VARCHAR(20),
            room_id INTEGER REFERENCES rooms(id)
        );
        
        CREATE TABLE IF NOT EXISTS rooms (
            id INTEGER PRIMARY KEY,
            etazh INTEGER NOT NULL,
            kol_mest INTEGER NOT NULL,
            svobodno INTEGER NOT NULL
        );
        
        CREATE TABLE IF NOT EXISTS requests (
            id SERIAL PRIMARY KEY,
            type VARCHAR(20) NOT NULL,
            status VARCHAR(20) NOT NULL,
            student_id INTEGER REFERENCES students(id),
            room_id INTEGER REFERENCES rooms(id)
        );
        """
        db.execute_query(create_tables_query)

        # Добавляем тестовые данные, если таблицы пусты
        db.execute_query(
            "INSERT INTO rooms (id, etazh, kol_mest, svobodno) VALUES (101, 1, 3, 3) ON CONFLICT (id) DO NOTHING"
        )
        db.execute_query(
            "INSERT INTO rooms (id, etazh, kol_mest, svobodno) VALUES (102, 1, 3, 3) ON CONFLICT (id) DO NOTHING"
        )

        window = MainWindow()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        logging.error(f"Ошибка инициализации базы данных: {e}")
        QMessageBox.critical(
            None,
            "Критическая ошибка",
            f"Не удалось подключиться к базе данных: {str(e)}",
        )
        sys.exit(1)
